from __future__ import annotations

import tempfile
from datetime import date, datetime
from pathlib import Path

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Response,
    UploadFile,
    status,
)
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.deps import get_current_user, require_role
from app.db.session import get_db
from app.models.mis import MisSubmission
from app.models.mis_template import MisTemplate
from app.models.user import User
from app.schemas.mis_template import (
    MisTemplateCreate,
    MisTemplateResponse,
    MisTemplateUpdate,
    TemplateCandidateRow,
    TemplateCandidatesResponse,
    TemplateDryRunResponse,
    TemplateDryRunRow,
)
from app.services.mis.template_runner import (
    TemplateRunError,
    run_template,
)

router = APIRouter(prefix="/mis/templates", tags=["mis-templates"])

_writer = require_role(["ADMIN", "ANALYST"])

_MAX_UPLOAD_BYTES = 25 * 1024 * 1024


def _to_response(tpl: MisTemplate) -> MisTemplateResponse:
    return MisTemplateResponse.model_validate(tpl)


def _save_temp_xlsx(content: bytes) -> Path:
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(content)
    tmp.flush()
    tmp.close()
    return Path(tmp.name)


def _clear_other_defaults(db: Session, tpl: MisTemplate) -> None:
    if not tpl.is_default:
        return
    db.execute(
        select(MisTemplate)
        .where(
            MisTemplate.company_id == tpl.company_id,
            MisTemplate.is_default.is_(True),
            MisTemplate.id != tpl.id,
        )
    )
    others = (
        db.execute(
            select(MisTemplate).where(
                MisTemplate.company_id == tpl.company_id,
                MisTemplate.is_default.is_(True),
                MisTemplate.id != tpl.id,
            )
        )
        .scalars()
        .all()
    )
    for other in others:
        other.is_default = False


@router.get("", response_model=list[MisTemplateResponse])
def list_templates(
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
    company_id: str | None = Query(None),
) -> list[MisTemplateResponse]:
    stmt = select(MisTemplate).order_by(MisTemplate.company_id.nulls_first(), MisTemplate.name)
    if company_id is not None:
        stmt = stmt.where(MisTemplate.company_id == company_id)
    rows = db.execute(stmt).scalars().all()
    return [_to_response(t) for t in rows]


@router.get("/{template_id}", response_model=MisTemplateResponse)
def get_template(
    template_id: int,
    db: Session = Depends(get_db),
    _user: User = Depends(get_current_user),
) -> MisTemplateResponse:
    tpl = db.get(MisTemplate, template_id)
    if tpl is None:
        raise HTTPException(status_code=404, detail="Template not found")
    return _to_response(tpl)


@router.post("", response_model=MisTemplateResponse, status_code=status.HTTP_201_CREATED)
def create_template(
    payload: MisTemplateCreate,
    db: Session = Depends(get_db),
    user: User = Depends(_writer),
) -> MisTemplateResponse:
    tpl = MisTemplate(
        company_id=payload.company_id,
        name=payload.name,
        version=1,
        is_default=payload.is_default,
        sheet_name_pattern=payload.sheet_name_pattern,
        header_row=payload.header_row,
        period_orientation=payload.period_orientation,
        row_mappings=[m.model_dump() for m in payload.row_mappings],
        created_by=user.id,
        updated_by=user.id,
    )
    db.add(tpl)
    db.flush()
    _clear_other_defaults(db, tpl)
    db.commit()
    db.refresh(tpl)
    return _to_response(tpl)


@router.patch("/{template_id}", response_model=MisTemplateResponse)
def update_template(
    template_id: int,
    payload: MisTemplateUpdate,
    db: Session = Depends(get_db),
    user: User = Depends(_writer),
) -> MisTemplateResponse:
    tpl = db.get(MisTemplate, template_id)
    if tpl is None:
        raise HTTPException(status_code=404, detail="Template not found")

    if payload.name is not None:
        tpl.name = payload.name
    if payload.sheet_name_pattern is not None:
        tpl.sheet_name_pattern = payload.sheet_name_pattern
    if payload.header_row is not None:
        tpl.header_row = payload.header_row
    if payload.period_orientation is not None:
        tpl.period_orientation = payload.period_orientation
    if payload.row_mappings is not None:
        tpl.row_mappings = [m.model_dump() for m in payload.row_mappings]
        tpl.version = (tpl.version or 1) + 1
    if payload.is_default is not None:
        tpl.is_default = payload.is_default

    tpl.updated_by = user.id
    db.flush()
    _clear_other_defaults(db, tpl)
    db.commit()
    db.refresh(tpl)
    return _to_response(tpl)


@router.post("/{template_id}/set-default", response_model=MisTemplateResponse)
def set_default_template(
    template_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(_writer),
) -> MisTemplateResponse:
    tpl = db.get(MisTemplate, template_id)
    if tpl is None:
        raise HTTPException(status_code=404, detail="Template not found")
    tpl.is_default = True
    tpl.updated_by = user.id
    db.flush()
    _clear_other_defaults(db, tpl)
    db.commit()
    db.refresh(tpl)
    return _to_response(tpl)


@router.delete(
    "/{template_id}",
    status_code=status.HTTP_204_NO_CONTENT,
    response_class=Response,
)
def delete_template(
    template_id: int,
    db: Session = Depends(get_db),
    _user: User = Depends(_writer),
) -> Response:
    tpl = db.get(MisTemplate, template_id)
    if tpl is None:
        raise HTTPException(status_code=404, detail="Template not found")
    in_use = db.execute(
        select(MisSubmission.id).where(MisSubmission.template_id == template_id).limit(1)
    ).first()
    if in_use is not None:
        raise HTTPException(
            status_code=409,
            detail="Template is referenced by one or more submissions; cannot delete.",
        )
    db.delete(tpl)
    db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.post("/extract-candidates", response_model=TemplateCandidatesResponse)
async def extract_candidates(
    file: UploadFile = File(...),
    sheet_name: str | None = Form(None),
    header_row: int = Form(1),
    label_col_index: int = Form(1),
    _user: User = Depends(_writer),
) -> TemplateCandidatesResponse:
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=415, detail="Only .xlsx files are supported")
    content = await file.read()
    if len(content) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File exceeds 25 MB cap")

    tmp_path = _save_temp_xlsx(content)
    try:
        try:
            wb = load_workbook(tmp_path, data_only=True, read_only=True)
        except Exception as exc:
            raise HTTPException(status_code=422, detail=f"Could not open workbook: {exc}") from exc

        try:
            sheet_names = list(wb.sheetnames)
            chosen = sheet_name or sheet_names[0]
            if chosen not in sheet_names:
                raise HTTPException(
                    status_code=400,
                    detail=f"Sheet {chosen!r} not in {sheet_names!r}",
                )
            ws = wb[chosen]
            rows = list(ws.iter_rows(values_only=True))
        finally:
            wb.close()

        header_idx = max(0, header_row - 1)
        period_cols: list[dict] = []
        if header_idx < len(rows):
            for col_idx, cell in enumerate(rows[header_idx]):
                if isinstance(cell, datetime):
                    period_cols.append(
                        {
                            "col_index": col_idx,
                            "month_date": date(cell.year, cell.month, 1).isoformat(),
                        }
                    )
                elif isinstance(cell, date):
                    period_cols.append(
                        {
                            "col_index": col_idx,
                            "month_date": date(cell.year, cell.month, 1).isoformat(),
                        }
                    )

        candidates: list[TemplateCandidateRow] = []
        for idx, row in enumerate(rows[header_idx + 1 :], start=header_idx + 1):
            if not row or label_col_index >= len(row):
                continue
            label = row[label_col_index]
            if not isinstance(label, str) or not label.strip():
                continue
            samples: list = []
            for pc in period_cols[:3]:
                ci = pc["col_index"]
                samples.append(row[ci] if ci < len(row) else None)
            candidates.append(
                TemplateCandidateRow(
                    row_index=idx,
                    label=label.strip(),
                    sample_values=[
                        str(v) if v is not None else None for v in samples
                    ],
                )
            )

        return TemplateCandidatesResponse(
            sheet_names=sheet_names,
            selected_sheet=chosen,
            header_row=header_row,
            period_columns=period_cols,
            rows=candidates,
        )
    finally:
        try:
            tmp_path.unlink()
        except OSError:
            pass


@router.post("/{template_id}/dry-run", response_model=TemplateDryRunResponse)
async def dry_run_template(
    template_id: int,
    file: UploadFile = File(...),
    company_id: str | None = Form(None),
    db: Session = Depends(get_db),
    _user: User = Depends(_writer),
) -> TemplateDryRunResponse:
    tpl = db.get(MisTemplate, template_id)
    if tpl is None:
        raise HTTPException(status_code=404, detail="Template not found")
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=415, detail="Only .xlsx files are supported")
    content = await file.read()
    if len(content) > _MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File exceeds 25 MB cap")

    # Detached copy so we can override company_id without persisting.
    tpl_copy = MisTemplate(
        company_id=company_id or tpl.company_id or "preview",
        name=tpl.name,
        version=tpl.version,
        sheet_name_pattern=tpl.sheet_name_pattern,
        header_row=tpl.header_row,
        period_orientation=tpl.period_orientation,
        row_mappings=tpl.row_mappings,
    )

    tmp_path = _save_temp_xlsx(content)
    try:
        try:
            parsed = run_template(tmp_path, tpl_copy)
        except TemplateRunError as exc:
            raise HTTPException(status_code=422, detail=str(exc)) from exc
        except NotImplementedError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
    finally:
        try:
            tmp_path.unlink()
        except OSError:
            pass

    sample = [
        TemplateDryRunRow(
            month_date=r.month_date.isoformat(),
            geography=r.geography,
            revenue_lacs=str(r.revenue_lacs) if r.revenue_lacs is not None else None,
            cogs_lacs=str(r.cogs_lacs) if r.cogs_lacs is not None else None,
            gross_margin_lacs=str(r.gross_margin_lacs)
            if r.gross_margin_lacs is not None
            else None,
            ebitda_lacs=str(r.ebitda_lacs) if r.ebitda_lacs is not None else None,
        )
        for r in parsed.monthly_rows[:8]
    ]

    return TemplateDryRunResponse(
        monthly_count=len(parsed.monthly_rows),
        bu_count=len(parsed.bu_rows),
        sample_monthly=sample,
        period_year=parsed.period_year,
        period_month=parsed.period_month,
    )
