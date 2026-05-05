"""Sprint 8 — Excel export builders. Generate workbooks against a seeded
in-memory DB, then re-open them with openpyxl to assert structure."""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.models.company import PortfolioCompany
from app.models.mis import MisBuMonthly, MisMonthly
from app.models.transaction import PortfolioTransaction
from app.services.exports.mis_export import (
    build_mis_bulk_workbook,
    build_mis_workbook,
)
from app.services.exports.portfolio_export import build_portfolio_workbook


def _seed_portfolio(db) -> None:
    co = PortfolioCompany(
        company_name="Acme",
        company_code="company_x",
        sector="Tech",
        portfolio_type="Entity_A",
        date_of_first_investment=date(2023, 4, 1),
        investment_value_cr=Decimal("-10.5"),
        current_value_cr=Decimal("18.0"),
        moic=Decimal("1.7142"),
        currency="INR",
    )
    db.add(co)
    db.commit()
    db.refresh(co)
    db.add(
        PortfolioTransaction(
            company_id=co.id,
            transaction_date=date(2023, 4, 1),
            transaction_type="Investment",
            amount_cr=Decimal("-10.5"),
            original_currency="INR",
            original_amount=Decimal("10.5"),
            amount_inr_cr=Decimal("-10.5"),
        )
    )
    db.commit()


def _seed_mis(db, company_id: str = "company_x") -> None:
    db.add(
        MisMonthly(
            company_id=company_id,
            month_date=date(2025, 6, 1),
            geography="consolidated",
            currency="INR",
            revenue_lacs=Decimal("420.0"),
            cogs_lacs=Decimal("180.0"),
            gross_margin_lacs=Decimal("240.0"),
            ebitda_lacs=Decimal("90.0"),
        )
    )
    db.add(
        MisBuMonthly(
            company_id=company_id,
            bu_id="BU_01",
            month_date=date(2025, 6, 1),
            currency="INR",
            revenue_lacs=Decimal("250.0"),
            ebitda_lacs=Decimal("60.0"),
        )
    )
    db.commit()


def test_portfolio_export_has_expected_sheets(db) -> None:
    _seed_portfolio(db)
    payload = build_portfolio_workbook(db)
    wb = load_workbook(BytesIO(payload), data_only=True)
    assert "Summary by sector" in wb.sheetnames
    assert "Companies" in wb.sheetnames
    assert "Transactions" in wb.sheetnames
    assert "Valuations" in wb.sheetnames
    # Companies sheet has a header row + the seeded company.
    ws = wb["Companies"]
    assert ws["A1"].value == "ID"
    assert ws.max_row >= 2


def test_mis_export_per_company(db) -> None:
    _seed_mis(db)
    payload = build_mis_workbook(db, "company_x")
    wb = load_workbook(BytesIO(payload), data_only=True)
    # Sheet name is the company code, capped to 31 chars.
    assert wb.sheetnames[0] == "company_x"
    ws = wb.active
    # First row is the header; second row is the seeded month.
    assert ws["A1"].value == "Month"
    assert ws.max_row >= 2


def test_mis_export_bulk_creates_one_sheet_per_company(db) -> None:
    _seed_mis(db, "company_a")
    _seed_mis(db, "company_b")
    payload = build_mis_bulk_workbook(db, ["company_a", "company_b"])
    wb = load_workbook(BytesIO(payload), data_only=True)
    assert set(wb.sheetnames) == {"company_a", "company_b"}
