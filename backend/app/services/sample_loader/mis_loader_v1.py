"""Parser for the canonical doc § 3.3 MIS layout (Company_01).

Produces structured rows for `mis_monthly`, `mis_bu_monthly`, and one wrapping
`mis_submissions`. Outlet-level rows are skipped for now — the layout is more
intricate (months × outlets × line items at varying offsets) and Sprint 5's
template builder is the proper home for that.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

# Canonical month columns on the Consolidated sheet are at indices 2..13.
CONSOLIDATED_MONTH_COL_RANGE = range(2, 14)

# BU sheet: months at indices 4..15.
BU_MONTH_COL_RANGE = range(4, 16)

# Channel rows on BU sheet → § 3.3 channel_*_lacs columns. Mapping is informed
# by the actual mock data (Channel_01 = dine-in flagship, Website = online, etc.).
CHANNEL_MAP_BU = {
    "Channel_01": "channel_dine_in_lacs",
    "Website": "channel_aggregator_a_lacs",
    "Events": "channel_catering_lacs",
    "B2B": "channel_aggregator_b_lacs",
    "Gifting": "channel_aggregator_d_lacs",
    "Channel_06": "channel_franchise_lacs",
}


@dataclass
class MisMonthlyRow:
    company_id: str
    month_date: date
    fiscal_year: str
    quarter: str
    geography: str
    currency: str = "INR"
    revenue_lacs: Decimal | None = None
    cogs_lacs: Decimal | None = None
    gross_margin_lacs: Decimal | None = None
    gross_margin_pct: Decimal | None = None
    total_operating_costs_lacs: Decimal | None = None
    ebitda_lacs: Decimal | None = None
    ebitda_pct: Decimal | None = None


@dataclass
class MisBuMonthlyRow:
    company_id: str
    bu_id: str
    month_date: date
    fiscal_year: str
    quarter: str
    currency: str = "INR"
    revenue_lacs: Decimal | None = None
    cogs_lacs: Decimal | None = None
    gross_margin_lacs: Decimal | None = None
    gross_margin_pct: Decimal | None = None
    operating_costs_lacs: Decimal | None = None
    ebitda_lacs: Decimal | None = None
    ebitda_pct: Decimal | None = None
    channel_dine_in_lacs: Decimal | None = None
    channel_aggregator_a_lacs: Decimal | None = None
    channel_aggregator_b_lacs: Decimal | None = None
    channel_aggregator_d_lacs: Decimal | None = None
    channel_catering_lacs: Decimal | None = None
    channel_franchise_lacs: Decimal | None = None


@dataclass
class ParsedMisSubmission:
    company_id: str
    period_year: int
    period_month: int
    fiscal_year: str
    source_file_name: str
    monthly_rows: list[MisMonthlyRow] = field(default_factory=list)
    bu_rows: list[MisBuMonthlyRow] = field(default_factory=list)


def _to_decimal(value: object) -> Decimal | None:
    if value is None or value == "" or value == "-" or (isinstance(value, str) and value.strip() in {"-", "na", "NA"}):
        return None
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _normalize_geography(label: str) -> str:
    """'Country_A ' → 'country_a', 'Country_A - Without ITC reversal' → 'country_a'."""
    base = label.strip().split(" - ")[0].strip()
    return base.lower()


def _fiscal_year_for(month_date: date) -> str:
    """Indian FY: April → March. April 2025 = FY26."""
    fy_end = month_date.year + (1 if month_date.month >= 4 else 0)
    return f"FY{str(fy_end)[-2:]}"


def _quarter_for(month_date: date) -> str:
    fy_month = (month_date.month - 4) % 12 + 1  # April = 1
    return f"Q{(fy_month - 1) // 3 + 1}"


def _section_label(label: str) -> str | None:
    """Map a col-B label to the canonical section it belongs to."""
    s = label.strip().lower()
    if s == "net revenue" or s.endswith("net revenue"):
        return "revenue"
    if s == "less: cogs" or s.endswith("cogs") and "%" not in s and "less" in s:
        return "cogs"
    if s in {"less: cogs"} or s == "less: cogs":
        return "cogs"
    if "gross margin %" in s:
        return "gross_margin_pct"
    if s.startswith("gross margin"):
        return "gross_margin"
    if s.startswith("less: operating costs"):
        return "operating_costs_in"
    if s.startswith("operating costs"):
        # Could be the "Operating costs" % section. Disambiguate by suffix.
        return "operating_costs"  # treated as primary numeric value section
    if "ebitda %" in s or "operating profit / (loss) %" in s.lower():
        return "ebitda_pct"
    if "ebitda" in s or "operating profit / (loss)" in s.lower():
        return "ebitda"
    return None


def _classify_b_label(raw: object) -> tuple[str, str] | None:
    """Returns (kind, geography) where kind is 'header'|'geo'|'total'.

    geo is one of: 'country_a', 'city_z', 'consolidated' (for total rows), or '' for headers.
    Returns None for unrecognized labels.
    """
    if not isinstance(raw, str):
        return None
    s = raw.strip()
    if not s:
        return None
    low = s.lower()
    if low.startswith("total "):
        return ("total", "consolidated")
    geo = _normalize_geography(s)
    if geo in {"country_a", "city_z"}:
        return ("geo", geo)
    return ("header", "")


def _load_consolidated(ws, company_id: str) -> list[MisMonthlyRow]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = rows[1]
    months: list[date | None] = []
    for col in CONSOLIDATED_MONTH_COL_RANGE:
        cell = header[col] if col < len(header) else None
        if isinstance(cell, datetime):
            # First-of-month for consistency.
            months.append(date(cell.year, cell.month, 1))
        else:
            months.append(None)

    # Build (geography, month_date) → MisMonthlyRow accumulator.
    acc: dict[tuple[str, date], MisMonthlyRow] = {}

    def get(geo: str, m: date) -> MisMonthlyRow:
        key = (geo, m)
        if key not in acc:
            acc[key] = MisMonthlyRow(
                company_id=company_id,
                month_date=m,
                fiscal_year=_fiscal_year_for(m),
                quarter=_quarter_for(m),
                geography=geo,
            )
        return acc[key]

    current_section: str | None = None

    for row in rows[2:]:
        if not row or len(row) < 2:
            continue
        label = row[1]
        cls = _classify_b_label(label)
        if cls is None:
            continue
        kind, geo = cls
        if kind == "header":
            section = _section_label(str(label))
            if section is not None:
                current_section = section
            continue
        if current_section is None:
            continue

        for slot, m in enumerate(months):
            if m is None:
                continue
            col = CONSOLIDATED_MONTH_COL_RANGE.start + slot
            value = row[col] if col < len(row) else None
            d = _to_decimal(value)
            if d is None:
                continue
            target = get(geo, m)
            if current_section == "revenue":
                target.revenue_lacs = d
            elif current_section == "cogs":
                target.cogs_lacs = d
            elif current_section == "gross_margin":
                target.gross_margin_lacs = d
            elif current_section == "gross_margin_pct":
                target.gross_margin_pct = d
            elif current_section == "operating_costs_in":
                target.total_operating_costs_lacs = d
            elif current_section == "operating_costs":
                # If we already filled total_operating_costs_lacs from "Less: Operating Costs"
                # section, the "Operating costs" % section follows; treat as redundant.
                if target.total_operating_costs_lacs is None:
                    target.total_operating_costs_lacs = d
            elif current_section == "ebitda":
                target.ebitda_lacs = d
            elif current_section == "ebitda_pct":
                target.ebitda_pct = d

    return list(acc.values())


def _load_bu(ws, company_id: str, bu_id: str) -> list[MisBuMonthlyRow]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = rows[1]
    months: list[date | None] = []
    for col in BU_MONTH_COL_RANGE:
        cell = header[col] if col < len(header) else None
        if isinstance(cell, datetime):
            months.append(date(cell.year, cell.month, 1))
        else:
            months.append(None)

    # Build month → MisBuMonthlyRow.
    acc: dict[date, MisBuMonthlyRow] = {}

    def get(m: date) -> MisBuMonthlyRow:
        if m not in acc:
            acc[m] = MisBuMonthlyRow(
                company_id=company_id,
                bu_id=bu_id,
                month_date=m,
                fiscal_year=_fiscal_year_for(m),
                quarter=_quarter_for(m),
            )
        return acc[m]

    def value_for_col(row: tuple, slot: int) -> Decimal | None:
        col = BU_MONTH_COL_RANGE.start + slot
        return _to_decimal(row[col] if col < len(row) else None)

    for row in rows:
        if not row or len(row) < 4:
            continue
        # Section labels live in col B (idx 1) — e.g. "TOTAL INCOME", "GROSS MARGIN", "INDIRECT COSTS".
        # Sub-section labels in col C (idx 2). Channel labels in col D (idx 3).
        b_label = row[1]
        c_label = row[2]
        d_label = row[3]

        # Top-level metrics (col B):
        if isinstance(b_label, str):
            top = b_label.strip().upper()
            if top == "TOTAL INCOME":
                for slot, m in enumerate(months):
                    if m is None:
                        continue
                    v = value_for_col(row, slot)
                    if v is not None:
                        get(m).revenue_lacs = v
                continue
            if top == "GROSS MARGIN":
                for slot, m in enumerate(months):
                    if m is None:
                        continue
                    v = value_for_col(row, slot)
                    if v is not None:
                        get(m).gross_margin_lacs = v
                continue
            if top == "INDIRECT COSTS":
                for slot, m in enumerate(months):
                    if m is None:
                        continue
                    v = value_for_col(row, slot)
                    if v is not None:
                        get(m).operating_costs_lacs = v
                continue

        # Sub-section (col C):
        if isinstance(c_label, str):
            sub = c_label.strip()
            if sub.upper() == "COGS":
                for slot, m in enumerate(months):
                    if m is None:
                        continue
                    v = value_for_col(row, slot)
                    if v is not None:
                        get(m).cogs_lacs = v
                continue
            if sub == "% GM":
                for slot, m in enumerate(months):
                    if m is None:
                        continue
                    v = value_for_col(row, slot)
                    if v is not None:
                        get(m).gross_margin_pct = v
                continue

        # Channel rows (col D under "Income from Operations"):
        if isinstance(d_label, str) and d_label.strip() in CHANNEL_MAP_BU:
            target_field = CHANNEL_MAP_BU[d_label.strip()]
            for slot, m in enumerate(months):
                if m is None:
                    continue
                v = value_for_col(row, slot)
                if v is not None:
                    setattr(get(m), target_field, v)
            continue

    # Compute EBITDA = revenue - cogs - operating_costs where all three are known.
    for m, r in acc.items():
        if r.revenue_lacs is not None and r.cogs_lacs is not None and r.operating_costs_lacs is not None:
            r.ebitda_lacs = r.revenue_lacs - r.cogs_lacs - r.operating_costs_lacs
            if r.revenue_lacs and r.revenue_lacs != 0:
                r.ebitda_pct = r.ebitda_lacs / r.revenue_lacs

    return list(acc.values())


def load_mis_v1(xlsx_path: Path, company_id: str) -> ParsedMisSubmission:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        consolidated_rows = _load_consolidated(wb["Consolidated MIS FY 2026"], company_id)
    except KeyError:
        consolidated_rows = []

    bu_rows: list[MisBuMonthlyRow] = []
    for bu_sheet, bu_id in [("BU_01 MIS", "BU_01"), ("BU_02 MIS", "BU_02")]:
        if bu_sheet in wb.sheetnames:
            bu_rows.extend(_load_bu(wb[bu_sheet], company_id, bu_id))

    wb.close()

    # Pick the latest month seen as the submission's headline period.
    all_months = [r.month_date for r in consolidated_rows] + [r.month_date for r in bu_rows]
    if all_months:
        latest = max(all_months)
    else:
        latest = date.today()

    return ParsedMisSubmission(
        company_id=company_id,
        period_year=latest.year,
        period_month=latest.month,
        fiscal_year=_fiscal_year_for(latest),
        source_file_name=xlsx_path.name,
        monthly_rows=consolidated_rows,
        bu_rows=bu_rows,
    )
