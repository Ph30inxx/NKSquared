"""Best-effort parser for the Company_02 MIS layout.

The Company_02 file is in a non-standard channel-by-channel layout. We extract
total monthly revenue (sum of every top-level sales row) and per-BU revenue.
COGS / EBITDA / channel splits stay NULL — Sprint 5's template builder is the
proper home for parsers like this. Doc § 11 anticipates this exact problem.
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.services.sample_loader.mis_loader_v1 import (
    MisBuMonthlyRow,
    MisMonthlyRow,
    ParsedMisSubmission,
    _fiscal_year_for,
    _quarter_for,
    _to_decimal,
)

# Header row r02 has month strings like "April -25", "May -25", … "March -26", "Total".
MONTH_NAME_TO_NUMBER = {
    "april": 4, "may": 5, "june": 6, "july": 7, "aug": 8, "august": 8,
    "sept": 9, "september": 9, "oct": 10, "october": 10, "nov": 11, "november": 11,
    "dec": 12, "december": 12, "jan": 1, "january": 1, "feb": 2, "february": 2,
    "march": 3, "mar": 3,
}


def _parse_month_header(s: str) -> date | None:
    """'April -25' → date(2025, 4, 1). 'Total' → None."""
    if not isinstance(s, str):
        return None
    s = s.strip().lower().rstrip(".")
    if s in {"total", "particulars", ""}:
        return None
    # Split on space and "-"
    parts = s.replace("-", " ").split()
    if len(parts) < 2:
        return None
    month = MONTH_NAME_TO_NUMBER.get(parts[0])
    if month is None:
        return None
    try:
        yy = int(parts[-1])
    except ValueError:
        return None
    year = 2000 + yy if yy < 100 else yy
    return date(year, month, 1)


def load_mis_v2(xlsx_path: Path, company_id: str) -> ParsedMisSubmission:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        ws = wb["MIS Report FY25-26"]
    except KeyError:
        wb.close()
        return ParsedMisSubmission(
            company_id=company_id,
            period_year=date.today().year,
            period_month=date.today().month,
            fiscal_year=_fiscal_year_for(date.today()),
            source_file_name=xlsx_path.name,
        )

    rows = list(ws.iter_rows(values_only=True))

    # Find the header row with month names.
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and len(row) > 2 and isinstance(row[1], str) and row[1].strip() == "Particulars":
            header_row_idx = i
            break
    if header_row_idx is None:
        wb.close()
        return ParsedMisSubmission(
            company_id=company_id,
            period_year=date.today().year,
            period_month=date.today().month,
            fiscal_year=_fiscal_year_for(date.today()),
            source_file_name=xlsx_path.name,
        )

    header = rows[header_row_idx]
    months: list[date | None] = [None]  # index 0 unused
    for col in range(1, len(header)):
        months.append(_parse_month_header(header[col]) if isinstance(header[col], str) else None)

    # Top-level revenue rows we'll sum into the company's monthly revenue. The file
    # mixes raw INR values with channel breakdowns — we want the unindented totals only.
    TOP_LEVEL_REVENUE_ROWS = {
        "Dine-in Sales (i)",
        "Catering Sales & Scrap Value (ii)",
        "Aggregator Aggregator_A Sales (iii) (Net of GST)",
        "Aggregator Aggregator_B Sales (iv) (Net of GST)",
        "Aggregator Aggregator_D Sales (v) (Net of GST)",
    }

    monthly: dict[date, MisMonthlyRow] = {}

    def m_get(m: date) -> MisMonthlyRow:
        if m not in monthly:
            monthly[m] = MisMonthlyRow(
                company_id=company_id,
                month_date=m,
                fiscal_year=_fiscal_year_for(m),
                quarter=_quarter_for(m),
                geography="consolidated",
                revenue_lacs=Decimal("0"),
            )
        return monthly[m]

    bu_rows: dict[tuple[str, date], MisBuMonthlyRow] = {}

    def bu_get(bu_id: str, m: date) -> MisBuMonthlyRow:
        key = (bu_id, m)
        if key not in bu_rows:
            bu_rows[key] = MisBuMonthlyRow(
                company_id=company_id,
                bu_id=bu_id,
                month_date=m,
                fiscal_year=_fiscal_year_for(m),
                quarter=_quarter_for(m),
                revenue_lacs=Decimal("0"),
            )
        return bu_rows[key]

    for row in rows[header_row_idx + 1:]:
        if not row or len(row) < 2:
            continue
        label = row[1]
        if not isinstance(label, str):
            continue
        label = label.strip()

        if label in TOP_LEVEL_REVENUE_ROWS:
            for slot, m in enumerate(months):
                if m is None:
                    continue
                v = _to_decimal(row[slot] if slot < len(row) else None)
                if v is None:
                    continue
                # Source amounts are in raw INR — convert to lacs (÷ 1e5).
                lacs = v / Decimal("100000")
                m_get(m).revenue_lacs = (m_get(m).revenue_lacs or Decimal("0")) + lacs
            continue

        # Per-BU rows: "BU_01" / "BU_02" / ... up to "BU_12". The file has these
        # repeatedly under each top-level sales channel. We aggregate across.
        if label.startswith("BU_") and len(label) <= 6:
            bu_id = label
            for slot, m in enumerate(months):
                if m is None:
                    continue
                v = _to_decimal(row[slot] if slot < len(row) else None)
                if v is None:
                    continue
                lacs = v / Decimal("100000")
                row_obj = bu_get(bu_id, m)
                row_obj.revenue_lacs = (row_obj.revenue_lacs or Decimal("0")) + lacs

    wb.close()

    monthly_rows = [r for r in monthly.values() if r.revenue_lacs and r.revenue_lacs > 0]
    bu_rows_list = [r for r in bu_rows.values() if r.revenue_lacs and r.revenue_lacs > 0]
    if monthly_rows:
        latest = max(r.month_date for r in monthly_rows)
    else:
        latest = date.today()

    return ParsedMisSubmission(
        company_id=company_id,
        period_year=latest.year,
        period_month=latest.month,
        fiscal_year=_fiscal_year_for(latest),
        source_file_name=xlsx_path.name,
        monthly_rows=monthly_rows,
        bu_rows=bu_rows_list,
    )
