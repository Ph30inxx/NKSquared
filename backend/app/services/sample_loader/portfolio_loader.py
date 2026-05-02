"""Parse the master portfolio xlsm into structured records ready for DB insert.

The workbook layout is documented in the loader plan; key features:
  - "Portfolio Master" lists every company with one row per company.
  - Each company has a per-sheet view named after the company in col C.
  - Per-sheet layout is "key in col B, consolidated value in col C, per-tranche
    values in cols E onward" (col D is empty).
  - Amounts on the Portfolio Master are in INR Cr; per-sheet amounts are in
    raw INR rupees (divide by 1e7 for Cr).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from app.services.sample_loader.mappings import (
    PORTFOLIO_TYPE_MAP,
    normalize_asset_class,
    normalize_investment_status,
    normalize_portfolio_status,
    normalize_portfolio_type,
)

# Per-sheet labels — looked up case-insensitively against col B. Some sheets
# (e.g. Strategic Equity ones like Company_45) have an extra "Entity_D Holding
# Post-issue" row that shifts everything by one, so we resolve by label, not
# by absolute row index.
LABEL_NUM_TRANCHES = "number of tranches"
LABEL_SERIES = "series"
LABEL_TXN_DATE = "latest date of investment"
LABEL_TXN_TYPE = "type of transaction"
LABEL_INSTRUMENT_TYPE = "type of instrument"
LABEL_INVESTING_ENTITY = "investing entity"
LABEL_SHARE_PRICE = "share price"
LABEL_ENTRY_PRE_MONEY = "entry pre money valuation"
LABEL_ENTRY_POST_MONEY = "entry post money valuation"
LABEL_INVESTMENT_SHAREHOLDING = "investment shareholding"
LABEL_TRANCHE_TOTAL_SHARES = "entity_d total shares"
LABEL_TOTAL_INVESTED_RAW = "total invested amount"
LABEL_LATEST_PRE_MONEY = "latest pre money valuation"
LABEL_LATEST_POST_MONEY = "latest post money valuation"
LABEL_LATEST_INVESTMENT_VALUATION_RAW = "latest investment valuation"

# Tranches start at column index 4 (col E); col D is intentionally empty.
TRANCHE_START_COL = 4

# Portfolio Master column indices (0-based).
PM_COL_PARTICULAR = 1
PM_COL_NAME = 2
PM_COL_COUNTRY = 3
PM_COL_CURRENCY = 4
PM_COL_PORTFOLIO_STATUS = 5
PM_COL_INVESTMENT_STATUS = 6
PM_COL_PORTFOLIO_TYPE = 7
PM_COL_ASSET_CLASS = 8
PM_COL_SECTOR = 9
PM_COL_SUB_SECTOR = 10
PM_COL_FIRST_INVESTMENT = 13
PM_COL_TOTAL_INVESTED_CR = 15
PM_COL_CURRENT_VALUE_CR = 16

DATA_START_ROW = 3  # rows 0..2 are headers / blank

# Section dividers in col B that we always skip.
DIVIDER_LABELS = frozenset({
    "Realised Portfolio",
    "Unrealised Portfolio",
})


@dataclass
class ParsedTransaction:
    transaction_date: date
    transaction_type: str  # "Investment" / "Follow_on"
    series: str | None
    instrument_type: str | None
    investing_entity: str | None
    raw_txn_type: str | None  # "Primary" / "Secondary" / "IPO ..." for notes
    amount_cr: Decimal  # always positive magnitude
    shares: Decimal | None
    share_price: Decimal | None
    pre_money_valuation_cr: Decimal | None
    post_money_valuation_cr: Decimal | None
    shareholding_pct: Decimal | None


@dataclass
class ParsedValuation:
    valuation_date: date
    pre_money_valuation_cr: Decimal | None
    post_money_valuation_cr: Decimal | None


@dataclass
class ParsedCompany:
    company_name: str               # canonical (matches sheet name)
    display_name: str | None
    portfolio_type: str | None
    investment_status: str | None
    portfolio_status: str | None
    asset_class: str | None
    sector: str | None
    sub_sector: str | None
    country: str | None
    date_of_first_investment: date | None
    current_value_cr: Decimal | None
    currency: str
    transactions: list[ParsedTransaction] = field(default_factory=list)
    valuation: ParsedValuation | None = None
    has_detail_sheet: bool = False


def _to_decimal(value: object) -> Decimal | None:
    if value is None or value == "" or value == "-":
        return None
    if isinstance(value, str):
        try:
            return Decimal(value)
        except Exception:
            return None
    try:
        return Decimal(str(value))
    except Exception:
        return None


def _to_cr(raw: object) -> Decimal | None:
    """Convert a raw INR amount to INR Cr."""
    d = _to_decimal(raw)
    if d is None:
        return None
    return d / Decimal("10000000")


def _to_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _is_aggregation_row(row: tuple) -> bool:
    """The summary block at the bottom of Portfolio Master uses portfolio_type
    labels in col B and Unrealized/Realized in col C — not real companies."""
    label = row[PM_COL_PARTICULAR] if len(row) > PM_COL_PARTICULAR else None
    name = row[PM_COL_NAME] if len(row) > PM_COL_NAME else None
    if not isinstance(label, str) or not isinstance(name, str):
        return False
    return label.strip() in PORTFOLIO_TYPE_MAP and name.strip() in {"Unrealized", "Realized"}


def _parse_master_row(row: tuple) -> dict | None:
    """Pull metadata from a Portfolio Master row. Returns None for non-company rows."""
    if len(row) <= PM_COL_NAME:
        return None
    particular = row[PM_COL_PARTICULAR]
    name = row[PM_COL_NAME]
    if particular is None and name is None:
        return None
    if isinstance(particular, str) and particular.strip() in DIVIDER_LABELS:
        return None
    if name is None:
        return None
    if _is_aggregation_row(row):
        return None

    company_name = str(name).strip()
    if not company_name:
        return None

    currency = row[PM_COL_CURRENCY] if len(row) > PM_COL_CURRENCY else None
    if not isinstance(currency, str) or currency.strip() in {"", "-"}:
        currency = "INR"
    else:
        currency = currency.strip().upper()

    return {
        "company_name": company_name,
        "display_name": str(particular).strip() if isinstance(particular, str) else None,
        "country": row[PM_COL_COUNTRY] if isinstance(row[PM_COL_COUNTRY], str) else None,
        "currency": currency,
        "portfolio_status": normalize_portfolio_status(row[PM_COL_PORTFOLIO_STATUS]),
        "investment_status": normalize_investment_status(row[PM_COL_INVESTMENT_STATUS]),
        "portfolio_type": normalize_portfolio_type(row[PM_COL_PORTFOLIO_TYPE]),
        "asset_class": normalize_asset_class(row[PM_COL_ASSET_CLASS]),
        "sector": row[PM_COL_SECTOR] if isinstance(row[PM_COL_SECTOR], str) else None,
        "sub_sector": row[PM_COL_SUB_SECTOR] if isinstance(row[PM_COL_SUB_SECTOR], str) else None,
        "date_of_first_investment": _to_date(row[PM_COL_FIRST_INVESTMENT]),
        "current_value_cr": _to_decimal(row[PM_COL_CURRENT_VALUE_CR]),
    }


def _build_label_index(rows: list[tuple]) -> dict[str, int]:
    """Map lowercased col-B label → row index. Earlier rows win on duplicates."""
    out: dict[str, int] = {}
    for idx, row in enumerate(rows):
        if len(row) < 2:
            continue
        b = row[1]
        if not isinstance(b, str):
            continue
        key = b.strip().lower()
        if key and key not in out:
            out[key] = idx
    return out


def _cell_at(rows: list[tuple], row_idx: int | None, col_idx: int) -> object:
    if row_idx is None or row_idx >= len(rows):
        return None
    row = rows[row_idx]
    if col_idx >= len(row):
        return None
    return row[col_idx]


def _clamp_pct(v: Decimal | None) -> Decimal | None:
    """shareholding_pct is NUMERIC(8,6) — only [0, 1] makes sense. Drop bad rows."""
    if v is None:
        return None
    if v < 0 or v > 1:
        return None
    return v


def _parse_company_sheet(rows: list[tuple]) -> tuple[list[ParsedTransaction], ParsedValuation | None, Decimal | None]:
    """Read tranches + latest valuation + latest current value from a per-company sheet."""
    labels = _build_label_index(rows)

    n_tranches_raw = _cell_at(rows, labels.get(LABEL_NUM_TRANCHES), 2)
    try:
        n_tranches = int(n_tranches_raw) if n_tranches_raw is not None else 0
    except (TypeError, ValueError):
        n_tranches = 0

    transactions: list[ParsedTransaction] = []
    for i in range(n_tranches):
        col = TRANCHE_START_COL + i
        txn_date = _to_date(_cell_at(rows, labels.get(LABEL_TXN_DATE), col))
        if txn_date is None:
            continue
        amount_raw = _cell_at(rows, labels.get(LABEL_TOTAL_INVESTED_RAW), col)
        amount_cr = _to_cr(amount_raw) or Decimal("0")
        if amount_cr <= 0:
            # No money means this isn't a real tranche — skip rather than insert a 0-row.
            continue
        raw_type = _cell_at(rows, labels.get(LABEL_TXN_TYPE), col)
        raw_type_str = str(raw_type).strip() if isinstance(raw_type, str) else None
        series_v = _cell_at(rows, labels.get(LABEL_SERIES), col)
        instrument_v = _cell_at(rows, labels.get(LABEL_INSTRUMENT_TYPE), col)
        entity_v = _cell_at(rows, labels.get(LABEL_INVESTING_ENTITY), col)
        transactions.append(
            ParsedTransaction(
                transaction_date=txn_date,
                transaction_type="",  # filled below after sort
                series=str(series_v).strip() if isinstance(series_v, str) else None,
                instrument_type=str(instrument_v).strip() if isinstance(instrument_v, str) else None,
                investing_entity=str(entity_v).strip() if isinstance(entity_v, str) else None,
                raw_txn_type=raw_type_str,
                amount_cr=amount_cr,
                shares=_to_decimal(_cell_at(rows, labels.get(LABEL_TRANCHE_TOTAL_SHARES), col)),
                share_price=_to_decimal(_cell_at(rows, labels.get(LABEL_SHARE_PRICE), col)),
                pre_money_valuation_cr=_to_cr(_cell_at(rows, labels.get(LABEL_ENTRY_PRE_MONEY), col)),
                post_money_valuation_cr=_to_cr(_cell_at(rows, labels.get(LABEL_ENTRY_POST_MONEY), col)),
                shareholding_pct=_clamp_pct(_to_decimal(_cell_at(rows, labels.get(LABEL_INVESTMENT_SHAREHOLDING), col))),
            )
        )

    # Sort by date; first → Investment, rest → Follow_on.
    transactions.sort(key=lambda t: t.transaction_date)
    for i, t in enumerate(transactions):
        t.transaction_type = "Investment" if i == 0 else "Follow_on"

    # Latest valuation row (col C is the consolidated value).
    latest_pre = _to_cr(_cell_at(rows, labels.get(LABEL_LATEST_PRE_MONEY), 2))
    latest_post = _to_cr(_cell_at(rows, labels.get(LABEL_LATEST_POST_MONEY), 2))
    latest_value_cr = _to_cr(_cell_at(rows, labels.get(LABEL_LATEST_INVESTMENT_VALUATION_RAW), 2))

    valuation: ParsedValuation | None = None
    if latest_post is not None and latest_post > 0:
        # Best date proxy: the most recent transaction; otherwise today.
        if transactions:
            valuation_date = max(t.transaction_date for t in transactions)
        else:
            valuation_date = date.today()
        valuation = ParsedValuation(
            valuation_date=valuation_date,
            pre_money_valuation_cr=latest_pre,
            post_money_valuation_cr=latest_post,
        )

    return transactions, valuation, latest_value_cr


def load_portfolio(xlsm_path: Path) -> list[ParsedCompany]:
    wb: Workbook = load_workbook(xlsm_path, data_only=True, read_only=True)
    try:
        master = wb["Portfolio Master"]
    except KeyError as e:
        raise ValueError(f"'Portfolio Master' sheet not found in {xlsm_path}") from e

    master_rows = list(master.iter_rows(values_only=True))
    sheet_names = set(wb.sheetnames)
    parsed: list[ParsedCompany] = []

    for i, row in enumerate(master_rows[DATA_START_ROW:], start=DATA_START_ROW):
        meta = _parse_master_row(row)
        if meta is None:
            continue

        company_name = meta["company_name"]

        # Locate the detail sheet — try col C (canonical), then col B (display).
        sheet_name: str | None = None
        if company_name in sheet_names:
            sheet_name = company_name
        elif meta["display_name"] and meta["display_name"] in sheet_names:
            sheet_name = meta["display_name"]

        transactions: list[ParsedTransaction] = []
        valuation: ParsedValuation | None = None
        current_value_cr: Decimal | None = meta["current_value_cr"]

        if sheet_name is not None:
            ws = wb[sheet_name]
            sheet_rows = list(ws.iter_rows(values_only=True))
            transactions, valuation, sheet_current = _parse_company_sheet(sheet_rows)
            # Prefer the per-sheet "Latest Investment Valuation" (more authoritative for some rows)
            # but fall back to Portfolio Master col Q.
            if sheet_current is not None and sheet_current > 0:
                current_value_cr = sheet_current

        parsed.append(
            ParsedCompany(
                company_name=company_name,
                display_name=meta["display_name"],
                portfolio_type=meta["portfolio_type"],
                investment_status=meta["investment_status"],
                portfolio_status=meta["portfolio_status"],
                asset_class=meta["asset_class"],
                sector=meta["sector"],
                sub_sector=meta["sub_sector"],
                country=meta["country"],
                date_of_first_investment=meta["date_of_first_investment"],
                current_value_cr=current_value_cr,
                currency=meta["currency"],
                transactions=transactions,
                valuation=valuation,
                has_detail_sheet=sheet_name is not None,
            )
        )

    wb.close()
    return parsed
