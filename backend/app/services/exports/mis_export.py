"""Per-company MIS Excel export. One sheet per company; the bulk variant
emits a single workbook with one sheet per company in a list."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.mis import MisBuMonthly, MisMonthly
from app.services.exports.workbook_builder import (
    autosize_columns,
    workbook_to_bytes,
    write_header,
    write_rows,
)


_MONTHLY_HEADERS = [
    "Month",
    "Geography",
    "Currency",
    "Revenue (Lacs)",
    "COGS (Lacs)",
    "Gross margin (Lacs)",
    "GM %",
    "Operating costs (Lacs)",
    "EBITDA (Lacs)",
    "EBITDA %",
    "Submission ID",
]

_BU_HEADERS = [
    "Month",
    "BU",
    "Currency",
    "Revenue (Lacs)",
    "COGS (Lacs)",
    "Gross margin (Lacs)",
    "GM %",
    "Operating costs (Lacs)",
    "EBITDA (Lacs)",
    "EBITDA %",
]


def _safe_sheet_name(raw: str) -> str:
    """Excel sheet names cap at 31 chars and forbid /\\?*[]:."""
    forbidden = set("/\\?*[]:")
    cleaned = "".join(c for c in raw if c not in forbidden) or "Sheet"
    return cleaned[:31]


def _write_company_sheet(ws: Worksheet, db: Session, company_id: str) -> None:
    write_header(ws, _MONTHLY_HEADERS)
    monthly = db.execute(
        select(MisMonthly)
        .where(MisMonthly.company_id == company_id)
        .order_by(MisMonthly.month_date.desc(), MisMonthly.geography)
    ).scalars()
    last_row = write_rows(
        ws,
        (
            [
                m.month_date,
                m.geography,
                m.currency,
                m.revenue_lacs,
                m.cogs_lacs,
                m.gross_margin_lacs,
                m.gross_margin_pct,
                m.total_operating_costs_lacs,
                m.ebitda_lacs,
                m.ebitda_pct,
                m.submission_id,
            ]
            for m in monthly
        ),
    )

    # BU rows below a blank row, with their own header.
    if last_row > 0:
        bu_header_row = last_row + 3
    else:
        bu_header_row = 4

    from copy import copy

    header_font = copy(ws.cell(row=1, column=1).font)
    for idx, h in enumerate(_BU_HEADERS, start=1):
        cell = ws.cell(row=bu_header_row, column=idx, value=h)
        cell.font = header_font

    bu_rows = db.execute(
        select(MisBuMonthly)
        .where(MisBuMonthly.company_id == company_id)
        .order_by(MisBuMonthly.month_date.desc(), MisBuMonthly.bu_id)
    ).scalars()
    for r_idx, b in enumerate(bu_rows, start=bu_header_row + 1):
        for c_idx, value in enumerate(
            [
                b.month_date,
                b.bu_id,
                b.currency,
                b.revenue_lacs,
                b.cogs_lacs,
                b.gross_margin_lacs,
                b.gross_margin_pct,
                b.operating_costs_lacs,
                b.ebitda_lacs,
                b.ebitda_pct,
            ],
            start=1,
        ):
            ws.cell(row=r_idx, column=c_idx, value=value)

    autosize_columns(ws)


def build_mis_workbook(db: Session, company_id: str) -> bytes:
    wb = Workbook()
    default = wb.active
    if default is not None:
        default.title = _safe_sheet_name(company_id)
        ws = default
    else:
        ws = wb.create_sheet(_safe_sheet_name(company_id))
    _write_company_sheet(ws, db, company_id)
    return workbook_to_bytes(wb)


def build_mis_bulk_workbook(db: Session, company_ids: list[str]) -> bytes:
    wb = Workbook()
    default = wb.active
    for idx, company_id in enumerate(company_ids):
        if idx == 0 and default is not None:
            default.title = _safe_sheet_name(company_id)
            ws = default
        else:
            ws = wb.create_sheet(_safe_sheet_name(company_id))
        _write_company_sheet(ws, db, company_id)
    if not company_ids and default is not None:
        default.title = "Empty"
    return workbook_to_bytes(wb)
