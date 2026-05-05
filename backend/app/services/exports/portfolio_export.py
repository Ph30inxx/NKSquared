"""Portfolio-wide Excel export: companies, transactions, valuations + summary."""
from __future__ import annotations

from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.company import PortfolioCompany
from app.models.transaction import PortfolioTransaction
from app.models.valuation import Valuation
from app.services import portfolio_service
from app.services.exports.workbook_builder import (
    autosize_columns,
    workbook_to_bytes,
    write_header,
    write_rows,
)


def _companies_sheet(wb: Workbook, db: Session) -> None:
    ws = wb.create_sheet("Companies")
    write_header(
        ws,
        [
            "ID",
            "Company name",
            "Code",
            "Sector",
            "Sub-sector",
            "Country",
            "Portfolio type",
            "Investment status",
            "Portfolio status",
            "First investment",
            "Invested (₹ Cr)",
            "Current value (₹ Cr)",
            "MOIC",
            "IRR",
            "Currency",
        ],
    )
    rows = db.execute(
        select(PortfolioCompany)
        .where(PortfolioCompany.is_active.is_(True))
        .order_by(PortfolioCompany.company_name)
    ).scalars()
    write_rows(
        ws,
        (
            [
                c.id,
                c.company_name,
                c.company_code,
                c.sector,
                c.sub_sector,
                c.country,
                c.portfolio_type,
                c.investment_status,
                c.portfolio_status,
                c.date_of_first_investment,
                abs(c.investment_value_cr) if c.investment_value_cr is not None else None,
                c.current_value_cr,
                c.moic,
                c.irr,
                c.currency,
            ]
            for c in rows
        ),
    )
    autosize_columns(ws)


def _transactions_sheet(wb: Workbook, db: Session) -> None:
    ws = wb.create_sheet("Transactions")
    write_header(
        ws,
        [
            "Date",
            "Company ID",
            "Type",
            "Amount (Cr, original)",
            "Currency",
            "Original amount",
            "Amount (₹ Cr)",
            "FX rate",
            "Series",
            "Instrument",
            "Investing entity",
            "Pre-money (Cr)",
            "Post-money (Cr)",
            "Shareholding %",
            "SSA reference",
            "Notes",
        ],
    )
    rows = db.execute(
        select(PortfolioTransaction).order_by(PortfolioTransaction.transaction_date.desc())
    ).scalars()
    write_rows(
        ws,
        (
            [
                t.transaction_date,
                t.company_id,
                t.transaction_type,
                t.amount_cr,
                t.original_currency,
                t.original_amount,
                t.amount_inr_cr,
                t.fx_rate_used,
                t.series,
                t.instrument_type,
                t.investing_entity,
                t.pre_money_valuation_cr,
                t.post_money_valuation_cr,
                t.shareholding_pct,
                t.ssa_reference,
                t.notes,
            ]
            for t in rows
        ),
    )
    autosize_columns(ws)


def _valuations_sheet(wb: Workbook, db: Session) -> None:
    ws = wb.create_sheet("Valuations")
    write_header(
        ws,
        [
            "Date",
            "Company ID",
            "Post-money (Cr)",
            "Pre-money (Cr)",
            "Currency",
            "Source",
            "Notes",
        ],
    )
    rows = db.execute(
        select(Valuation).order_by(Valuation.valuation_date.desc())
    ).scalars()
    write_rows(
        ws,
        (
            [
                v.valuation_date,
                v.company_id,
                v.post_money_valuation_cr,
                v.pre_money_valuation_cr,
                v.currency,
                v.source,
                v.notes,
            ]
            for v in rows
        ),
    )
    autosize_columns(ws)


def _summary_sheet(wb: Workbook, db: Session) -> None:
    ws = wb.create_sheet("Summary by sector", index=0)
    write_header(
        ws,
        ["Sector", "Companies", "Invested (₹ Cr)", "Current value (₹ Cr)", "MOIC"],
    )
    write_rows(
        ws,
        (
            [b.key, b.count, b.invested_cr, b.current_cr, b.weighted_moic]
            for b in portfolio_service.by_sector(db)
        ),
    )
    autosize_columns(ws)


def build_portfolio_workbook(db: Session) -> bytes:
    wb = Workbook()
    # openpyxl seeds a default 'Sheet'; remove it after we've added our own.
    default = wb.active
    _summary_sheet(wb, db)
    _companies_sheet(wb, db)
    _transactions_sheet(wb, db)
    _valuations_sheet(wb, db)
    if default is not None and default.title == "Sheet":
        del wb["Sheet"]
    return workbook_to_bytes(wb)
