"""Tiny helpers wrapping openpyxl. Centralized so every export sheet
shares header styling, freeze panes, and number formats."""
from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")

NUMBER_FORMAT = "#,##0.00"
PCT_FORMAT = "0.00%"
DATE_FORMAT = "yyyy-mm-dd"


def write_header(ws: Worksheet, headers: list[str]) -> None:
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
    ws.freeze_panes = "A2"


def write_rows(ws: Worksheet, rows: Iterable[list]) -> int:
    count = 0
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            if isinstance(value, Decimal):
                ws.cell(row=r_idx, column=c_idx, value=float(value))
            else:
                ws.cell(row=r_idx, column=c_idx, value=value)
        count += 1
    return count


def autosize_columns(ws: Worksheet, max_width: int = 40) -> None:
    if ws.max_row == 0:
        return
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        longest = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for v in row:
                if v is None:
                    continue
                length = len(str(v))
                if length > longest:
                    longest = length
        ws.column_dimensions[letter].width = min(max(longest + 2, 10), max_width)


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
