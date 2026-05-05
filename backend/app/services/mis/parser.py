"""MIS parser dispatch. Resolves which template to apply, then delegates to
either the legacy hardcoded loaders (`v1`/`v2`) or the generic
`template_runner` driven by a stored `MisTemplate.row_mappings` spec.

Resolution order:
  1. Explicit `template_id` argument → that template (DB-backed).
  2. Default template for the submission's company_id (DB-backed).
  3. Legacy heuristic — sheet-name match for the original v1/v2 layouts.
"""
from __future__ import annotations

from pathlib import Path
from typing import Literal

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.mis_template import MisTemplate
from app.services.mis.template_runner import (
    TemplateRunError,
    run_template,
)
from app.services.sample_loader.mis_loader_v1 import (
    ParsedMisSubmission,
    load_mis_v1,
)
from app.services.sample_loader.mis_loader_v2 import load_mis_v2

TemplateName = str  # 'v1', 'v2', or 'template:<id>'

V1_SHEET = "Consolidated MIS FY 2026"
V2_SHEET = "MIS Report FY25-26"


class UnknownTemplateError(ValueError):
    """Raised when a workbook matches none of the known MIS templates."""


def detect_template(file_path: Path) -> Literal["v1", "v2"] | None:
    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:
        raise UnknownTemplateError(
            f"Could not open workbook (not a valid .xlsx): {exc}"
        ) from exc
    try:
        names = set(wb.sheetnames)
    finally:
        wb.close()
    if V1_SHEET in names:
        return "v1"
    if V2_SHEET in names:
        return "v2"
    return None


def _load_db_template(db: Session, template_id: int) -> MisTemplate:
    tpl = db.execute(
        select(MisTemplate).where(MisTemplate.id == template_id)
    ).scalar_one_or_none()
    if tpl is None:
        raise UnknownTemplateError(f"Template id={template_id} not found")
    return tpl


def _default_template(db: Session, company_id: str) -> MisTemplate | None:
    return db.execute(
        select(MisTemplate).where(
            MisTemplate.company_id == company_id, MisTemplate.is_default.is_(True)
        )
    ).scalar_one_or_none()


def parse(
    file_path: Path,
    *,
    company_id: str,
    template_id: int | None = None,
    db: Session | None = None,
) -> tuple[TemplateName, ParsedMisSubmission]:
    if template_id is not None:
        if db is None:
            raise UnknownTemplateError("template_id requires a DB session")
        tpl = _load_db_template(db, template_id)
        # Override company so the parsed rows attach to the submission's company.
        tpl.company_id = company_id
        try:
            parsed = run_template(file_path, tpl)
        except TemplateRunError as exc:
            raise UnknownTemplateError(str(exc)) from exc
        return f"template:{tpl.id}", parsed

    if db is not None:
        tpl = _default_template(db, company_id)
        if tpl is not None:
            tpl.company_id = company_id
            try:
                parsed = run_template(file_path, tpl)
                return f"template:{tpl.id}", parsed
            except TemplateRunError:
                # Fall through to legacy heuristic if the saved template can't
                # parse this particular file (e.g. analyst uploaded a different
                # company's workbook by accident).
                pass

    legacy = detect_template(file_path)
    if legacy is None:
        raise UnknownTemplateError(
            f"Template not recognized. Expected sheet {V1_SHEET!r} or {V2_SHEET!r}, "
            "or configure a template under MIS → Templates."
        )
    if legacy == "v1":
        return legacy, load_mis_v1(file_path, company_id=company_id)
    return legacy, load_mis_v2(file_path, company_id=company_id)
