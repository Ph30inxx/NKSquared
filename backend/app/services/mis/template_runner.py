"""Generic Excel parser driven by a `MisTemplate.row_mappings` JSON spec.

The visual template builder produces a list of mapping entries; this module
applies them to a workbook and emits the same `ParsedMisSubmission` shape that
the legacy `mis_loader_v1`/`mis_loader_v2` parsers produce, so downstream
preview/approve flows do not branch on template source.

Mapping entry shape (each item in `row_mappings`):
    {
      "label_regex": "^Net Revenue$",     # matched against col B / configurable column
      "metric_code": "revenue_lacs",      # column on MisMonthly or MisBuMonthly
      "geography": "consolidated",         # optional; defaults to consolidated
      "bu_id": null,                       # if set, row populates MisBuMonthly instead
      "label_col_index": 1                 # optional, 0-based; default 1
    }

Period orientation `columns` means months run across the header row and metrics
down — the common case. `rows` (months down, metrics across) is not yet
implemented and raises `NotImplementedError`.
"""
from __future__ import annotations

import re
from dataclasses import asdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

from app.models.mis_template import MisTemplate
from app.services.sample_loader.mis_loader_v1 import (
    MisBuMonthlyRow,
    MisMonthlyRow,
    ParsedMisSubmission,
    _fiscal_year_for,
    _quarter_for,
)


_MONTHLY_COLS = {c for c in MisMonthlyRow.__dataclass_fields__}
_BU_COLS = {c for c in MisBuMonthlyRow.__dataclass_fields__}


class TemplateRunError(ValueError):
    """Raised when a workbook cannot be parsed against the supplied template."""


def _to_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        s = value.strip()
        if not s or s in {"-", "na", "NA", "N/A"}:
            return None
        try:
            return Decimal(s.replace(",", ""))
        except InvalidOperation:
            return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _pick_sheet(wb, pattern: str | None) -> str:
    if not pattern:
        return wb.sheetnames[0]
    rx = re.compile(pattern)
    for name in wb.sheetnames:
        if rx.search(name):
            return name
    raise TemplateRunError(
        f"No sheet matched template pattern {pattern!r}; sheets={wb.sheetnames!r}"
    )


def _extract_period_columns(header_row: tuple) -> dict[int, date]:
    """Return {col_index: month_date} for every cell in header_row that holds a date."""
    out: dict[int, date] = {}
    for idx, cell in enumerate(header_row):
        if isinstance(cell, datetime):
            out[idx] = date(cell.year, cell.month, 1)
        elif isinstance(cell, date):
            out[idx] = date(cell.year, cell.month, 1)
    return out


def run_template(file_path: Path, template: MisTemplate) -> ParsedMisSubmission:
    if template.period_orientation != "columns":
        raise NotImplementedError(
            f"period_orientation={template.period_orientation!r} not supported yet"
        )

    company_id = template.company_id or ""
    if not company_id:
        raise TemplateRunError(
            "Template runner requires a company-scoped template (company_id is NULL)"
        )

    try:
        wb = load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:
        raise TemplateRunError(f"Could not open workbook: {exc}") from exc

    try:
        sheet_name = _pick_sheet(wb, template.sheet_name_pattern)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if not rows:
        raise TemplateRunError(f"Sheet {sheet_name!r} is empty")

    header_idx = max(0, template.header_row - 1)
    if header_idx >= len(rows):
        raise TemplateRunError(
            f"header_row={template.header_row} but sheet has only {len(rows)} rows"
        )
    period_cols = _extract_period_columns(rows[header_idx])
    if not period_cols:
        raise TemplateRunError(
            f"No date cells found in header row {template.header_row}; "
            "check header_row and that period columns contain real Excel dates"
        )

    monthly_acc: dict[tuple[str, date], MisMonthlyRow] = {}
    bu_acc: dict[tuple[str, date], MisBuMonthlyRow] = {}

    compiled = []
    for entry in template.row_mappings or []:
        if not isinstance(entry, dict):
            continue
        if entry.get("_legacy"):
            # Sentinel for the seeded v1/v2 templates — they aren't drivable by
            # the generic runner; the legacy loaders handle them instead.
            raise TemplateRunError(
                "Legacy template marker encountered; use the legacy loader path"
            )
        label_regex = entry.get("label_regex")
        metric_code = entry.get("metric_code")
        if not label_regex or not metric_code:
            continue
        try:
            rx = re.compile(label_regex, re.IGNORECASE)
        except re.error:
            continue
        compiled.append(
            (
                rx,
                metric_code,
                entry.get("geography") or "consolidated",
                entry.get("bu_id"),
                int(entry.get("label_col_index", 1)),
            )
        )

    if not compiled:
        raise TemplateRunError("Template has no usable row_mappings")

    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        for rx, metric_code, geography, bu_id, label_col in compiled:
            if label_col >= len(row):
                continue
            label = row[label_col]
            if not isinstance(label, str) or not rx.search(label):
                continue

            for col_idx, month_date in period_cols.items():
                if col_idx >= len(row):
                    continue
                val = _to_decimal(row[col_idx])
                if val is None:
                    continue

                if bu_id:
                    if metric_code not in _BU_COLS:
                        continue
                    key = (bu_id, month_date)
                    target = bu_acc.get(key)
                    if target is None:
                        target = MisBuMonthlyRow(
                            company_id=company_id,
                            bu_id=bu_id,
                            month_date=month_date,
                            fiscal_year=_fiscal_year_for(month_date),
                            quarter=_quarter_for(month_date),
                        )
                        bu_acc[key] = target
                    setattr(target, metric_code, val)
                else:
                    if metric_code not in _MONTHLY_COLS:
                        continue
                    key = (geography, month_date)
                    target_m = monthly_acc.get(key)
                    if target_m is None:
                        target_m = MisMonthlyRow(
                            company_id=company_id,
                            month_date=month_date,
                            fiscal_year=_fiscal_year_for(month_date),
                            quarter=_quarter_for(month_date),
                            geography=geography,
                        )
                        monthly_acc[key] = target_m
                    setattr(target_m, metric_code, val)
            break  # one mapping per row is enough

    monthly_rows = list(monthly_acc.values())
    bu_rows = list(bu_acc.values())

    all_months = [r.month_date for r in monthly_rows] + [r.month_date for r in bu_rows]
    latest = max(all_months) if all_months else date.today()

    return ParsedMisSubmission(
        company_id=company_id,
        period_year=latest.year,
        period_month=latest.month,
        fiscal_year=_fiscal_year_for(latest),
        source_file_name=file_path.name,
        monthly_rows=monthly_rows,
        bu_rows=bu_rows,
    )


def parsed_to_dict(parsed: ParsedMisSubmission) -> dict:
    """JSON-safe rendering used for `mis_submissions.last_parse_payload` cache."""
    return {
        "company_id": parsed.company_id,
        "period_year": parsed.period_year,
        "period_month": parsed.period_month,
        "fiscal_year": parsed.fiscal_year,
        "source_file_name": parsed.source_file_name,
        "monthly_rows": [_dataclass_to_jsonable(r) for r in parsed.monthly_rows],
        "bu_rows": [_dataclass_to_jsonable(r) for r in parsed.bu_rows],
    }


def _dataclass_to_jsonable(row) -> dict:
    out: dict = {}
    for k, v in asdict(row).items():
        if isinstance(v, Decimal):
            out[k] = str(v)
        elif isinstance(v, (date, datetime)):
            out[k] = v.isoformat()
        else:
            out[k] = v
    return out
